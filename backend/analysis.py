import math

import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import plotly.express as px
import plotly.io as pio

def clean_columns(df):
    df.columns = (
        df.columns
        .str.strip()
        .str.replace(",", "", regex=False)
        .str.upper()
    )

    df.rename(columns={
        "ITEM ARCTILE": "ITEM ARTICLE",
        "CATERGORY": "CATEGORY"
    }, inplace=True)

    return df


def analyze_files(purchase_file, sales_file, retailer_info=None):
    if retailer_info is None:
        retailer_info = {}

    purchase_df = pd.read_excel(purchase_file)
    sales_df = pd.read_excel(sales_file)

    purchase_df = clean_columns(purchase_df)
    sales_df = clean_columns(sales_df)

    purchase_required = [
        "ITEM ARTICLE", "QTY", "SIZE", "PURCHASE PRICE",
        "COLOUR", "TOTAL AMT", "CATEGORY", "FIT", "MRP"
    ]

    sales_required = [
        "ITEM NAME", "QTY", "MRP", "DISCOUNT AMOUNT",
        "COLOUR", "NET AMOUNT", "CATEGORY", "MARGIN",
        "BILL NO", "DATE", "TAX AMOUNT", "SIZE"
    ]

    for col in purchase_required:
        if col not in purchase_df.columns:
            raise Exception(f"Purchase file missing column: {col}")

    for col in sales_required:
        if col not in sales_df.columns:
            raise Exception(f"Sales file missing column: {col}")

    purchase_df["QTY"] = pd.to_numeric(purchase_df["QTY"], errors="coerce").fillna(0)
    purchase_df["PURCHASE PRICE"] = pd.to_numeric(purchase_df["PURCHASE PRICE"], errors="coerce").fillna(0)
    purchase_df["MRP"] = pd.to_numeric(purchase_df["MRP"], errors="coerce").fillna(0)

    sales_df["QTY"] = pd.to_numeric(sales_df["QTY"], errors="coerce").fillna(0)
    sales_df["NET AMOUNT"] = pd.to_numeric(sales_df["NET AMOUNT"], errors="coerce").fillna(0)

    purchase_df["KEY"] = (
        purchase_df["ITEM ARTICLE"].astype(str).str.upper().str.strip()
        + "_"
        + purchase_df["COLOUR"].astype(str).str.upper().str.strip()
        + "_"
        + purchase_df["SIZE"].astype(str).str.upper().str.strip()
    )

    sales_df["KEY"] = (
        sales_df["ITEM NAME"].astype(str).str.upper().str.strip()
        + "_"
        + sales_df["COLOUR"].astype(str).str.upper().str.strip()
        + "_"
        + sales_df["SIZE"].astype(str).str.upper().str.strip()
    )

    purchase_summary = purchase_df.groupby("KEY").agg({
        "ITEM ARTICLE": "first",
        "QTY": "sum",
        "PURCHASE PRICE": "mean",
        "MRP": "mean",
        "COLOUR": "first",
        "SIZE": "first",
        "CATEGORY": "first",
        "FIT": "first"
    }).reset_index()

    purchase_summary.rename(columns={
        "ITEM ARTICLE": "ITEM NAME",
        "QTY": "PURCHASE QTY"
    }, inplace=True)

    sales_summary = sales_df.groupby("KEY").agg({
        "QTY": "sum",
        "NET AMOUNT": "sum"
    }).reset_index()

    sales_summary.rename(columns={
        "QTY": "SOLD QTY",
        "NET AMOUNT": "TOTAL SALES AMOUNT"
    }, inplace=True)

    final_df = pd.merge(purchase_summary, sales_summary, on="KEY", how="left")

    final_df["SOLD QTY"] = final_df["SOLD QTY"].fillna(0)
    final_df["TOTAL SALES AMOUNT"] = final_df["TOTAL SALES AMOUNT"].fillna(0)

    final_df["REMAINING STOCK"] = final_df["PURCHASE QTY"] - final_df["SOLD QTY"]

    final_df["SALES PERCENTAGE"] = (
        final_df["SOLD QTY"] / final_df["PURCHASE QTY"] * 100
    ).replace([float("inf"), -float("inf")], 0).fillna(0).round(2)

    final_df["PURCHASE COST OF SOLD ITEMS"] = final_df["SOLD QTY"] * final_df["PURCHASE PRICE"]

    final_df["PROFIT"] = (
        final_df["TOTAL SALES AMOUNT"] - final_df["PURCHASE COST OF SOLD ITEMS"]
    ).round(2)

    final_df["MARGIN PERCENTAGE"] = (
        final_df["PROFIT"] / final_df["TOTAL SALES AMOUNT"] * 100
    ).replace([float("inf"), -float("inf")], 0).fillna(0).round(2)

    def value_category(row):
        if row["SALES PERCENTAGE"] >= 70 and row["MARGIN PERCENTAGE"] >= 25:
            return "High Value"
        elif row["SALES PERCENTAGE"] >= 40:
            return "Moderate Value"
        else:
            return "Low Value"

    def reorder_decision(row):
        if row["SALES PERCENTAGE"] >= 70 and row["MARGIN PERCENTAGE"] >= 25:
            return "Reorder Immediately"
        elif row["SALES PERCENTAGE"] >= 40:
            return "Order Limited Quantity"
        else:
            return "Do Not Reorder"

    final_df["VALUE CATEGORY"] = final_df.apply(value_category, axis=1)
    final_df["REORDER DECISION"] = final_df.apply(reorder_decision, axis=1)

    final_df = final_df[[
        "ITEM NAME",
        "CATEGORY",
        "FIT",
        "COLOUR",
        "SIZE",
        "PURCHASE QTY",
        "SOLD QTY",
        "REMAINING STOCK",
        "PURCHASE PRICE",
        "MRP",
        "TOTAL SALES AMOUNT",
        "PROFIT",
        "MARGIN PERCENTAGE",
        "SALES PERCENTAGE",
        "VALUE CATEGORY",
        "REORDER DECISION"
    ]]

    report_folder = os.path.join("backend", "reports")
    os.makedirs(report_folder, exist_ok=True)

    # Excel report path
    output_path = os.path.join(report_folder, "final_analysis_report.xlsx")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        final_df.to_excel(writer, sheet_name="Item Analysis", index=False)

    # Format only Excel file
    format_professional_report(output_path, final_df, retailer_info)

    # Create separate HTML dashboard
    dashboard_path = os.path.join(report_folder, "stockpulse_dashboard.html")
    create_html_dashboard(dashboard_path, final_df, retailer_info)

    return output_path, final_df


def format_professional_report(output_path, final_df, retailer_info):
    wb = load_workbook(output_path)

    if "Summary" in wb.sheetnames:
        del wb["Summary"]

    analysis_ws = wb["Item Analysis"]
    summary_ws = wb.create_sheet("Summary", 0)

    # Colors
    navy = "0F172A"
    blue = "2563EB"
    light_blue = "DBEAFE"
    green = "DCFCE7"
    green_dark = "15803D"
    yellow = "FEF9C3"
    yellow_dark = "A16207"
    red = "FEE2E2"
    red_dark = "B91C1C"
    light = "F8FAFC"
    gray = "E2E8F0"
    white = "FFFFFF"

    dark_fill = PatternFill("solid", fgColor=navy)
    blue_fill = PatternFill("solid", fgColor=blue)
    light_blue_fill = PatternFill("solid", fgColor=light_blue)
    green_fill = PatternFill("solid", fgColor=green)
    yellow_fill = PatternFill("solid", fgColor=yellow)
    red_fill = PatternFill("solid", fgColor=red)
    light_fill = PatternFill("solid", fgColor=light)

    white_font = Font(color=white, bold=True)
    title_font = Font(size=20, bold=True, color=white)
    subtitle_font = Font(size=11, color="CBD5E1")
    section_font = Font(size=14, bold=True, color=navy)
    bold_font = Font(bold=True, color=navy)

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )

    # Sheet setup
    summary_ws.sheet_view.showGridLines = False
    analysis_ws.sheet_view.showGridLines = False

    for col in range(1, 11):
        summary_ws.column_dimensions[get_column_letter(col)].width = 18

    # Title banner
    summary_ws.merge_cells("A1:J2")
    summary_ws["A1"] = "StockPulse AI - Retail Garment Analysis Report"
    summary_ws["A1"].fill = dark_fill
    summary_ws["A1"].font = title_font
    summary_ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    summary_ws.merge_cells("A3:J3")
    summary_ws["A3"] = "Smart reorder recommendation report generated from purchase bill and sales report"
    summary_ws["A3"].fill = dark_fill
    summary_ws["A3"].font = subtitle_font
    summary_ws["A3"].alignment = Alignment(horizontal="center")

    summary_ws.row_dimensions[1].height = 32
    summary_ws.row_dimensions[2].height = 28
    summary_ws.row_dimensions[3].height = 24

    # Retailer details card
    summary_ws.merge_cells("A5:D5")
    summary_ws["A5"] = "Retailer Details"
    summary_ws["A5"].fill = blue_fill
    summary_ws["A5"].font = white_font
    summary_ws["A5"].alignment = Alignment(horizontal="center")

    retailer_rows = [
        ("Shop Name", retailer_info.get("shop_name", "")),
        ("Owner Name", retailer_info.get("owner_name", "")),
        ("Email", retailer_info.get("email", "")),
        ("Phone", retailer_info.get("phone", "")),
        ("City", retailer_info.get("city", "")),
        ("Generated On", datetime.now().strftime("%d-%m-%Y %I:%M %p")),
    ]

    start_row = 6
    for i, (label, value) in enumerate(retailer_rows):
        r = start_row + i
        summary_ws[f"A{r}"] = label
        summary_ws[f"B{r}"] = value
        summary_ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)

        summary_ws[f"A{r}"].font = bold_font
        summary_ws[f"A{r}"].fill = light_blue_fill
        summary_ws[f"B{r}"].fill = light_fill

        for col in range(1, 5):
            cell = summary_ws.cell(row=r, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # KPI calculation
    total_items = len(final_df)
    high_value = int((final_df["VALUE CATEGORY"] == "High Value").sum())
    moderate_value = int((final_df["VALUE CATEGORY"] == "Moderate Value").sum())
    low_value = int((final_df["VALUE CATEGORY"] == "Low Value").sum())

    total_purchase_qty = int(final_df["PURCHASE QTY"].sum())
    total_sold_qty = int(final_df["SOLD QTY"].sum())
    total_remaining = int(final_df["REMAINING STOCK"].sum())
    total_profit = float(final_df["PROFIT"].sum())

    # KPI cards
    kpi_cards = [
        ("Total Items", total_items, "A13:B15", light_blue_fill, blue),
        ("Sold Qty", total_sold_qty, "C13:D15", green_fill, green_dark),
        ("Remaining Stock", total_remaining, "E13:F15", yellow_fill, yellow_dark),
        ("Total Profit", f"₹{round(total_profit, 2)}", "G13:H15", green_fill, green_dark),
        ("Purchase Qty", total_purchase_qty, "I13:J15", light_blue_fill, blue),
    ]

    for title, value, cell_range, fill, font_color in kpi_cards:
        summary_ws.merge_cells(cell_range)
        top_left = cell_range.split(":")[0]
        cell = summary_ws[top_left]
        cell.value = f"{title}\n{value}"
        cell.fill = fill
        cell.font = Font(size=14, bold=True, color=font_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in summary_ws[cell_range]:
            for c in row:
                c.border = thin_border

    # Category summary table
    summary_ws.merge_cells("A18:D18")
    summary_ws["A18"] = "Value Category Summary"
    summary_ws["A18"].fill = dark_fill
    summary_ws["A18"].font = white_font
    summary_ws["A18"].alignment = Alignment(horizontal="center")

    category_data = [
        ("High Value", high_value, "Reorder Immediately", green_fill),
        ("Moderate Value", moderate_value, "Order Limited Quantity", yellow_fill),
        ("Low Value", low_value, "Do Not Reorder", red_fill),
    ]

    headers = ["Category", "Items", "Decision", "Meaning"]
    for col, header in enumerate(headers, start=1):
        cell = summary_ws.cell(row=19, column=col)
        cell.value = header
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    meanings = {
        "High Value": "Fast selling and profitable",
        "Moderate Value": "Average performance",
        "Low Value": "Slow moving or weak margin",
    }

    row = 20
    for category, count, decision, fill in category_data:
        values = [category, count, decision, meanings[category]]
        for col, value in enumerate(values, start=1):
            cell = summary_ws.cell(row=row, column=col)
            cell.value = value
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row += 1

    # Chart
    summary_ws["F18"] = "Category"
    summary_ws["G18"] = "Count"
    summary_ws["F19"] = "High Value"
    summary_ws["G19"] = high_value
    summary_ws["F20"] = "Moderate Value"
    summary_ws["G20"] = moderate_value
    summary_ws["F21"] = "Low Value"
    summary_ws["G21"] = low_value

    for cell in summary_ws["F18:G18"][0]:
        cell.fill = blue_fill
        cell.font = white_font
        cell.border = thin_border

    for r in range(19, 22):
        summary_ws[f"F{r}"].border = thin_border
        summary_ws[f"G{r}"].border = thin_border

    chart = BarChart()
    chart.title = "Item Value Category"
    chart.y_axis.title = "Item Count"
    chart.x_axis.title = "Category"
    chart.height = 7
    chart.width = 12

    data = Reference(summary_ws, min_col=7, min_row=18, max_row=21)
    cats = Reference(summary_ws, min_col=6, min_row=19, max_row=21)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    summary_ws.add_chart(chart, "F23")

    # Top reorder table
    summary_ws.merge_cells("A25:J25")
    summary_ws["A25"] = "Top Reorder Recommendations"
    summary_ws["A25"].fill = dark_fill
    summary_ws["A25"].font = white_font
    summary_ws["A25"].alignment = Alignment(horizontal="center")

    top_items = final_df.sort_values(
        by=["SALES PERCENTAGE", "MARGIN PERCENTAGE"],
        ascending=False
    ).head(8)

    top_headers = [
        "Item Name", "Category", "Colour", "Size", "Purchase Qty",
        "Sold Qty", "Remaining", "Sales %", "Value", "Decision"
    ]

    for col, header in enumerate(top_headers, start=1):
        cell = summary_ws.cell(row=26, column=col)
        cell.value = header
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    row = 27
    for _, item in top_items.iterrows():
        values = [
            item["ITEM NAME"],
            item["CATEGORY"],
            item["COLOUR"],
            item["SIZE"],
            item["PURCHASE QTY"],
            item["SOLD QTY"],
            item["REMAINING STOCK"],
            item["SALES PERCENTAGE"],
            item["VALUE CATEGORY"],
            item["REORDER DECISION"],
        ]

        for col, value in enumerate(values, start=1):
            cell = summary_ws.cell(row=row, column=col)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            if item["VALUE CATEGORY"] == "High Value":
                cell.fill = green_fill
            elif item["VALUE CATEGORY"] == "Moderate Value":
                cell.fill = yellow_fill
            else:
                cell.fill = red_fill

        row += 1

    # Format Item Analysis sheet
    analysis_ws.freeze_panes = "A2"
    analysis_ws.auto_filter.ref = analysis_ws.dimensions

    for cell in analysis_ws[1]:
        cell.fill = dark_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    header_map = {}
    for idx, cell in enumerate(analysis_ws[1], start=1):
        header_map[cell.value] = idx

    category_col = header_map.get("VALUE CATEGORY")
    decision_col = header_map.get("REORDER DECISION")

    for row in range(2, analysis_ws.max_row + 1):
        category_value = analysis_ws.cell(row=row, column=category_col).value

        if category_value == "High Value":
            fill = green_fill
        elif category_value == "Moderate Value":
            fill = yellow_fill
        else:
            fill = red_fill

        analysis_ws.cell(row=row, column=category_col).fill = fill
        analysis_ws.cell(row=row, column=decision_col).fill = fill

        for col in range(1, analysis_ws.max_column + 1):
            analysis_ws.cell(row=row, column=col).border = thin_border
            analysis_ws.cell(row=row, column=col).alignment = Alignment(vertical="center")

    money_cols = ["PURCHASE PRICE", "MRP", "TOTAL SALES AMOUNT", "PROFIT"]
    percent_cols = ["MARGIN PERCENTAGE", "SALES PERCENTAGE"]

    for col_name in money_cols:
        col_index = header_map.get(col_name)
        if col_index:
            for row in range(2, analysis_ws.max_row + 1):
                analysis_ws.cell(row=row, column=col_index).number_format = "₹#,##0.00"

    for col_name in percent_cols:
        col_index = header_map.get(col_name)
        if col_index:
            for row in range(2, analysis_ws.max_row + 1):
                analysis_ws.cell(row=row, column=col_index).number_format = "0.00"

    for col in range(1, analysis_ws.max_column + 1):
        column_letter = get_column_letter(col)
        max_length = 0

        for cell in analysis_ws[column_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        analysis_ws.column_dimensions[column_letter].width = min(max_length + 4, 30)

    # Better row heights
    for r in range(1, 40):
        summary_ws.row_dimensions[r].height = 24

    summary_ws.row_dimensions[13].height = 42
    summary_ws.row_dimensions[14].height = 42
    summary_ws.row_dimensions[15].height = 42

    wb.save(output_path)
def format_professional_report(output_path, final_df, retailer_info):
    wb = load_workbook(output_path)

    # Remove Summary sheet if it exists
    if "Summary" in wb.sheetnames:
        del wb["Summary"]

    ws = wb["Item Analysis"]
    ws.title = "Inventory Sheet"

    # Insert rows at top for retailer details
    ws.insert_rows(1, amount=9)

    # Colors
    navy = "0F172A"
    blue = "2563EB"
    light_blue = "DBEAFE"
    green = "DCFCE7"
    yellow = "FEF9C3"
    red = "FEE2E2"
    light = "F8FAFC"
    white = "FFFFFF"

    dark_fill = PatternFill("solid", fgColor=navy)
    blue_fill = PatternFill("solid", fgColor=blue)
    light_blue_fill = PatternFill("solid", fgColor=light_blue)
    green_fill = PatternFill("solid", fgColor=green)
    yellow_fill = PatternFill("solid", fgColor=yellow)
    red_fill = PatternFill("solid", fgColor=red)
    light_fill = PatternFill("solid", fgColor=light)

    white_font = Font(color=white, bold=True)
    title_font = Font(size=18, bold=True, color=white)
    subtitle_font = Font(size=11, color="DBEAFE")
    bold_font = Font(bold=True, color=navy)

    border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )

    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:P1")
    ws["A1"] = "StockPulse AI - Inventory Analysis Report"
    ws["A1"].fill = dark_fill
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:P2")
    ws["A2"] = "Retail Garment Sales, Stock and Reorder Decision Report"
    ws["A2"].fill = dark_fill
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # Retailer details heading
    ws.merge_cells("A4:P4")
    ws["A4"] = "Retailer / User Details"
    ws["A4"].fill = blue_fill
    ws["A4"].font = white_font
    ws["A4"].alignment = Alignment(horizontal="center", vertical="center")

    # Retailer details
    details = [
        ("Shop Name", retailer_info.get("shop_name", "")),
        ("Owner Name", retailer_info.get("owner_name", "")),
        ("Email", retailer_info.get("email", "")),
        ("Phone", retailer_info.get("phone", "")),
        ("City", retailer_info.get("city", "")),
        ("Generated On", datetime.now().strftime("%d-%m-%Y %I:%M %p")),
    ]

    positions = [
        ("A5", "B5"),
        ("D5", "E5"),
        ("G5", "H5"),
        ("A6", "B6"),
        ("D6", "E6"),
        ("G6", "H6"),
    ]

    for (label, value), (label_cell, value_cell) in zip(details, positions):
        ws[label_cell] = label
        ws[value_cell] = value

        ws[label_cell].font = bold_font
        ws[label_cell].fill = light_blue_fill
        ws[value_cell].fill = light_fill

        ws[label_cell].border = border
        ws[value_cell].border = border

        ws[label_cell].alignment = Alignment(horizontal="center", vertical="center")
        ws[value_cell].alignment = Alignment(horizontal="center", vertical="center")

    # Inventory table heading
    ws.merge_cells("A8:P8")
    ws["A8"] = "Inventory Analysis Sheet"
    ws["A8"].fill = dark_fill
    ws["A8"].font = white_font
    ws["A8"].alignment = Alignment(horizontal="center", vertical="center")

    # Header row is now row 10 because 9 rows were inserted
    header_row = 10
    data_start_row = 11

    ws.freeze_panes = "A11"
    ws.auto_filter.ref = f"A{header_row}:P{ws.max_row}"

    # Format table header
    for cell in ws[header_row]:
        cell.fill = dark_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Header map
    header_map = {}
    for idx, cell in enumerate(ws[header_row], start=1):
        header_map[cell.value] = idx

    category_col = header_map.get("VALUE CATEGORY")
    decision_col = header_map.get("REORDER DECISION")

    # Format data rows
    for row in range(data_start_row, ws.max_row + 1):
        category_value = ws.cell(row=row, column=category_col).value

        if category_value == "High Value":
            fill = green_fill
        elif category_value == "Moderate Value":
            fill = yellow_fill
        else:
            fill = red_fill

        if category_col:
            ws.cell(row=row, column=category_col).fill = fill

        if decision_col:
            ws.cell(row=row, column=decision_col).fill = fill

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Money columns
    money_cols = ["PURCHASE PRICE", "MRP", "TOTAL SALES AMOUNT", "PROFIT"]

    for col_name in money_cols:
        col_index = header_map.get(col_name)
        if col_index:
            for row in range(data_start_row, ws.max_row + 1):
                ws.cell(row=row, column=col_index).number_format = "₹#,##0.00"

    # Percentage columns
    percent_cols = ["MARGIN PERCENTAGE", "SALES PERCENTAGE"]

    for col_name in percent_cols:
        col_index = header_map.get(col_name)
        if col_index:
            for row in range(data_start_row, ws.max_row + 1):
                ws.cell(row=row, column=col_index).number_format = "0.00"

    # Column widths
    for col in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col)
        max_length = 0

        for cell in ws[column_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = min(max_length + 4, 28)

    # Row heights
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[4].height = 24
    ws.row_dimensions[8].height = 24
    ws.row_dimensions[10].height = 35

    wb.save(output_path)
def create_html_dashboard(output_path, final_df, retailer_info):
    dashboard_df = final_df.copy()

    # Suggested reorder quantity logic
    dashboard_df["SUGGESTED REORDER QTY"] = dashboard_df.apply(
        lambda row: max(
            0,
            math.ceil((row["SOLD QTY"] * 1.2) - row["REMAINING STOCK"])
        )
        if row["VALUE CATEGORY"] in ["High Value", "Moderate Value"]
        else 0,
        axis=1
    )

    # Priority logic
    def priority(row):
        if row["VALUE CATEGORY"] == "High Value":
            return "High Priority"
        elif row["VALUE CATEGORY"] == "Moderate Value":
            return "Medium Priority"
        else:
            return "Low Priority"

    dashboard_df["PRIORITY"] = dashboard_df.apply(priority, axis=1)

    # KPI values
    total_items = len(dashboard_df)
    total_purchase_qty = int(dashboard_df["PURCHASE QTY"].sum())
    total_sold_qty = int(dashboard_df["SOLD QTY"].sum())
    total_remaining = int(dashboard_df["REMAINING STOCK"].sum())
    total_profit = round(float(dashboard_df["PROFIT"].sum()), 2)

    high_value = int((dashboard_df["VALUE CATEGORY"] == "High Value").sum())
    moderate_value = int((dashboard_df["VALUE CATEGORY"] == "Moderate Value").sum())
    low_value = int((dashboard_df["VALUE CATEGORY"] == "Low Value").sum())

    reorder_required = int((dashboard_df["SUGGESTED REORDER QTY"] > 0).sum())
    slow_moving = int((dashboard_df["VALUE CATEGORY"] == "Low Value").sum())

    # Category analysis
    category_sales = (
        dashboard_df.groupby("CATEGORY")["TOTAL SALES AMOUNT"]
        .sum()
        .sort_values(ascending=False)
        .reset_index()
    )

    category_profit = (
        dashboard_df.groupby("CATEGORY")["PROFIT"]
        .sum()
        .sort_values(ascending=False)
        .reset_index()
    )

    best_category = category_sales.iloc[0]["CATEGORY"] if len(category_sales) > 0 else "N/A"
    best_category_sales = round(float(category_sales.iloc[0]["TOTAL SALES AMOUNT"]), 2) if len(category_sales) > 0 else 0

    best_profit_category = category_profit.iloc[0]["CATEGORY"] if len(category_profit) > 0 else "N/A"
    best_profit_value = round(float(category_profit.iloc[0]["PROFIT"]), 2) if len(category_profit) > 0 else 0

    top_item = dashboard_df.sort_values(
        by=["SALES PERCENTAGE", "MARGIN PERCENTAGE"],
        ascending=False
    ).head(1)

    best_item_name = top_item.iloc[0]["ITEM NAME"] if len(top_item) > 0 else "N/A"

    slow_items_df = dashboard_df.sort_values(
        by=["SALES PERCENTAGE", "SOLD QTY"],
        ascending=True
    ).head(10)

    # Charts
    value_count = dashboard_df["VALUE CATEGORY"].value_counts().reset_index()
    value_count.columns = ["VALUE CATEGORY", "COUNT"]

    fig1 = px.bar(
        value_count,
        x="VALUE CATEGORY",
        y="COUNT",
        text="COUNT",
        color="VALUE CATEGORY",
        title="Item Value Classification",
        color_discrete_map={
            "High Value": "#16a34a",
            "Moderate Value": "#f59e0b",
            "Low Value": "#dc2626"
        }
    )
    fig1.update_layout(template="plotly_white", title_x=0.5)

    fig2 = px.bar(
        category_sales.head(10),
        x="CATEGORY",
        y="TOTAL SALES AMOUNT",
        text="TOTAL SALES AMOUNT",
        title="Top Categories by Sales Amount"
    )
    fig2.update_layout(template="plotly_white", title_x=0.5)

    fig3 = px.bar(
        category_profit.head(10),
        x="CATEGORY",
        y="PROFIT",
        text="PROFIT",
        title="Top Categories by Profit"
    )
    fig3.update_layout(template="plotly_white", title_x=0.5)

    top_sold = dashboard_df.sort_values("SOLD QTY", ascending=False).head(10)
    fig4 = px.bar(
        top_sold,
        x="SOLD QTY",
        y="ITEM NAME",
        orientation="h",
        text="SOLD QTY",
        title="Top 10 Sold Items"
    )
    fig4.update_layout(template="plotly_white", title_x=0.5, yaxis={"categoryorder": "total ascending"})

    fig5 = px.scatter(
        dashboard_df,
        x="SALES PERCENTAGE",
        y="MARGIN PERCENTAGE",
        size="SOLD QTY",
        color="VALUE CATEGORY",
        hover_data=["ITEM NAME", "CATEGORY", "COLOUR", "SIZE"],
        title="Margin % vs Sales % Decision Map",
        color_discrete_map={
            "High Value": "#16a34a",
            "Moderate Value": "#f59e0b",
            "Low Value": "#dc2626"
        }
    )
    fig5.update_layout(template="plotly_white", title_x=0.5)

    stock_compare = dashboard_df.sort_values("REMAINING STOCK", ascending=False).head(10)
    fig6 = px.bar(
        stock_compare,
        x="ITEM NAME",
        y=["SOLD QTY", "REMAINING STOCK"],
        barmode="group",
        title="Sold Qty vs Remaining Stock"
    )
    fig6.update_layout(template="plotly_white", title_x=0.5)

    chart1 = pio.to_html(fig1, full_html=False, include_plotlyjs="cdn")
    chart2 = pio.to_html(fig2, full_html=False, include_plotlyjs=False)
    chart3 = pio.to_html(fig3, full_html=False, include_plotlyjs=False)
    chart4 = pio.to_html(fig4, full_html=False, include_plotlyjs=False)
    chart5 = pio.to_html(fig5, full_html=False, include_plotlyjs=False)
    chart6 = pio.to_html(fig6, full_html=False, include_plotlyjs=False)

    # Smart insights
    insights = [
        f"{high_value} items are classified as High Value and should be considered for reorder.",
        f"{low_value} items are Low Value and may need discounting or should not be reordered now.",
        f"{best_category} generated the highest sales amount of ₹{best_category_sales}.",
        f"{best_profit_category} generated the highest profit of ₹{best_profit_value}.",
        f"{best_item_name} is currently one of the best performing items.",
        f"{reorder_required} items have suggested reorder quantity greater than zero."
    ]

    insight_html = ""
    for item in insights:
        insight_html += f"<li>{item}</li>"

    # Reorder priority table
    reorder_table = dashboard_df.sort_values(
        by=["PRIORITY", "SALES PERCENTAGE", "MARGIN PERCENTAGE"],
        ascending=[True, False, False]
    ).head(20)

    reorder_rows = ""
    for _, row in reorder_table.iterrows():
        category_class = "low"
        if row["VALUE CATEGORY"] == "High Value":
            category_class = "high"
        elif row["VALUE CATEGORY"] == "Moderate Value":
            category_class = "moderate"

        reorder_rows += f"""
        <tr>
            <td>{row['PRIORITY']}</td>
            <td>{row['ITEM NAME']}</td>
            <td>{row['CATEGORY']}</td>
            <td>{row['COLOUR']}</td>
            <td>{row['SIZE']}</td>
            <td>{row['SOLD QTY']}</td>
            <td>{row['REMAINING STOCK']}</td>
            <td>{row['SALES PERCENTAGE']}%</td>
            <td>{row['MARGIN PERCENTAGE']}%</td>
            <td>{row['SUGGESTED REORDER QTY']}</td>
            <td><span class="badge {category_class}">{row['VALUE CATEGORY']}</span></td>
            <td>{row['REORDER DECISION']}</td>
        </tr>
        """

    # Slow moving table
    slow_rows = ""
    for _, row in slow_items_df.iterrows():
        slow_rows += f"""
        <tr>
            <td>{row['ITEM NAME']}</td>
            <td>{row['CATEGORY']}</td>
            <td>{row['COLOUR']}</td>
            <td>{row['SIZE']}</td>
            <td>{row['PURCHASE QTY']}</td>
            <td>{row['SOLD QTY']}</td>
            <td>{row['REMAINING STOCK']}</td>
            <td>{row['SALES PERCENTAGE']}%</td>
            <td>{row['REORDER DECISION']}</td>
        </tr>
        """

    shop_name = retailer_info.get("shop_name", "")
    owner_name = retailer_info.get("owner_name", "")
    email = retailer_info.get("email", "")
    phone = retailer_info.get("phone", "")
    city = retailer_info.get("city", "")

    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Retail Flow AI Dashboard</title>
        <style>
            body {{
                margin: 0;
                font-family: Arial, sans-serif;
                background: #f3f7ff;
                color: #111827;
            }}

            .header {{
                background: linear-gradient(135deg, #07111f, #0f2b5f, #2563eb);
                color: white;
                padding: 36px 55px;
            }}

            .header h1 {{
                margin: 0;
                font-size: 38px;
            }}

            .header p {{
                color: #dbeafe;
                margin-top: 8px;
            }}

            .container {{
                padding: 28px 55px 55px;
            }}

            .section {{
                background: white;
                padding: 24px;
                border-radius: 22px;
                box-shadow: 0 14px 35px rgba(15,23,42,0.08);
                margin-bottom: 26px;
                border: 1px solid #dbeafe;
            }}

            .retailer-grid,
            .kpi-grid {{
                display: grid;
                grid-template-columns: repeat(4, 1fr);
                gap: 16px;
            }}

            .info-box,
            .kpi {{
                background: #f8fbff;
                padding: 18px;
                border-radius: 18px;
                border: 1px solid #dbeafe;
            }}

            .info-box span,
            .kpi span {{
                display: block;
                color: #64748b;
                font-weight: bold;
                font-size: 13px;
            }}

            .info-box strong,
            .kpi strong {{
                display: block;
                margin-top: 8px;
                font-size: 25px;
                color: #2563eb;
            }}

            .insights {{
                background: linear-gradient(135deg, #eff6ff, #ffffff);
                border-left: 6px solid #2563eb;
            }}

            .insights ul {{
                margin: 0;
                padding-left: 20px;
                line-height: 1.9;
                color: #334155;
                font-weight: 600;
            }}

            .charts {{
                display: grid;
                grid-template-columns: 1fr 1fr;
                gap: 22px;
            }}

            .chart-card {{
                background: white;
                padding: 18px;
                border-radius: 22px;
                box-shadow: 0 14px 35px rgba(15,23,42,0.08);
                border: 1px solid #dbeafe;
            }}

            table {{
                width: 100%;
                border-collapse: collapse;
                min-width: 1100px;
            }}

            .table-wrap {{
                overflow-x: auto;
            }}

            th {{
                background: #111827;
                color: white;
                padding: 13px;
                font-size: 13px;
                text-align: left;
            }}

            td {{
                border-bottom: 1px solid #e5e7eb;
                padding: 12px;
                font-size: 14px;
            }}

            tr:nth-child(even) {{
                background: #f8fbff;
            }}

            .badge {{
                padding: 6px 10px;
                border-radius: 999px;
                font-weight: bold;
                font-size: 12px;
            }}

            .high {{
                background: #dcfce7;
                color: #166534;
            }}

            .moderate {{
                background: #fef9c3;
                color: #854d0e;
            }}

            .low {{
                background: #fee2e2;
                color: #991b1b;
            }}

            @media(max-width: 900px) {{
                .retailer-grid,
                .kpi-grid,
                .charts {{
                    grid-template-columns: 1fr;
                }}

                .container,
                .header {{
                    padding-left: 18px;
                    padding-right: 18px;
                }}
            }}
        </style>
    </head>

    <body>
        <div class="header">
            <h1>Retail Flow AI Dashboard</h1>
            <p>Smart garment sales analysis, inventory performance and reorder decision dashboard</p>
        </div>

        <div class="container">
            <div class="section">
                <h2>Retailer Details</h2>
                <div class="retailer-grid">
                    <div class="info-box"><span>Shop Name</span><strong>{shop_name}</strong></div>
                    <div class="info-box"><span>Owner</span><strong>{owner_name}</strong></div>
                    <div class="info-box"><span>Email</span><strong>{email}</strong></div>
                    <div class="info-box"><span>Phone</span><strong>{phone}</strong></div>
                    <div class="info-box"><span>City</span><strong>{city}</strong></div>
                </div>
            </div>

            <div class="section insights">
                <h2>Smart Insights</h2>
                <ul>
                    {insight_html}
                </ul>
            </div>

            <div class="section">
                <h2>Business KPIs</h2>
                <div class="kpi-grid">
                    <div class="kpi"><span>Total Items</span><strong>{total_items}</strong></div>
                    <div class="kpi"><span>Purchase Qty</span><strong>{total_purchase_qty}</strong></div>
                    <div class="kpi"><span>Sold Qty</span><strong>{total_sold_qty}</strong></div>
                    <div class="kpi"><span>Remaining Stock</span><strong>{total_remaining}</strong></div>
                    <div class="kpi"><span>Total Profit</span><strong>₹{total_profit}</strong></div>
                    <div class="kpi"><span>High Value Items</span><strong>{high_value}</strong></div>
                    <div class="kpi"><span>Low Value Items</span><strong>{low_value}</strong></div>
                    <div class="kpi"><span>Reorder Required</span><strong>{reorder_required}</strong></div>
                </div>
            </div>

            <div class="charts">
                <div class="chart-card">{chart1}</div>
                <div class="chart-card">{chart2}</div>
                <div class="chart-card">{chart3}</div>
                <div class="chart-card">{chart4}</div>
                <div class="chart-card">{chart5}</div>
                <div class="chart-card">{chart6}</div>
            </div>

            <div class="section">
                <h2>Reorder Priority Table</h2>
                <div class="table-wrap">
                    <table>
                        <thead>
                            <tr>
                                <th>Priority</th>
                                <th>Item</th>
                                <th>Category</th>
                                <th>Colour</th>
                                <th>Size</th>
                                <th>Sold Qty</th>
                                <th>Remaining</th>
                                <th>Sales %</th>
                                <th>Margin %</th>
                                <th>Suggested Reorder</th>
                                <th>Value</th>
                                <th>Decision</th>
                            </tr>
                        </thead>
                        <tbody>
                            {reorder_rows}
                        </tbody>
                    </table>
                </div>
            </div>

            <div class="section">
                <h2>Slow Moving Items</h2>
                <div class="table-wrap">
                    <table>
                        <thead>
                            <tr>
                                <th>Item</th>
                                <th>Category</th>
                                <th>Colour</th>
                                <th>Size</th>
                                <th>Purchase Qty</th>
                                <th>Sold Qty</th>
                                <th>Remaining</th>
                                <th>Sales %</th>
                                <th>Decision</th>
                            </tr>
                        </thead>
                        <tbody>
                            {slow_rows}
                        </tbody>
                    </table>
                </div>
            </div>
        </div>
    </body>
    </html>
    """

    with open(output_path, "w", encoding="utf-8") as file:
        file.write(html_content)
    